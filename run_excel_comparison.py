"""
Generate two Excel files for short vs long dataset comparison.

Usage:
    py -3.11 run_excel_comparison.py

Output:
    XAUUSD-Deep-Analysis/deep_short.xlsx   (3.5 months, 2026-01~04)
    XAUUSD-Deep-Analysis/deep_long.xlsx    (7 years, 2019~2026)

Sheets per file (identical structure, same row layout for easy comparison):
    策略概覽       -- win_rate / PF / net_pnl / max_dd / time_bleed% for all 3 strategies
    S2B垂頭陷阱    -- BB%B high/low group breakdown + fail-type distribution
    S1進場清單     -- by_4h_state + 4H x BB cross-table + by_dxy
    S2A進場清單    -- by_4h_state + by_dxy
    S2B進場清單    -- by_4h_state + by_dxy
    TimeBleed      -- tb_rate by 4H state / by session for all 3 strategies
"""
import matplotlib
matplotlib.use("Agg")

import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from analysis.config import get_strategies, DATASET_MODES, PRICE_CSV, PRICE_CSV_4H, XAUUSD_CSV_1D, DXY_CSV_1D
from analysis import loader, metrics
from analysis.deep_analysis import s2b_trap_analysis, entry_checklist, time_bleed_features

OUT_DIR = Path(__file__).parent / "XAUUSD-Deep-Analysis"
OUT_DIR.mkdir(exist_ok=True)

SID_ORDER = ["S1-AweWithBB", "S2A-RSI", "S2B-Hammer"]


# ── Data loading ──────────────────────────────────────────────────────────────

def _load_common_data() -> dict:
    return {
        "price_30m": loader.load_price(PRICE_CSV),
        "price_4h":  loader.load_price(PRICE_CSV_4H) if PRICE_CSV_4H.exists() else None,
        "price_1d":  loader.load_price(XAUUSD_CSV_1D) if XAUUSD_CSV_1D.exists() else None,
        "dxy_1d":    loader.load_dxy(DXY_CSV_1D) if DXY_CSV_1D.exists() else None,
    }


def _run_mode(mode: str, data: dict) -> dict:
    strats = get_strategies(mode)
    s_map = {s["id"]: s for s in strats}

    trades, summs = {}, {}
    for sid, cfg in s_map.items():
        t = loader.load_trades(cfg["folder"] / cfg["trades_csv"])
        trades[sid] = t
        s = metrics.summary(t)
        summs[sid] = {**s, "max_drawdown": metrics.max_drawdown(t)}

    p30 = data["price_30m"]
    p4h = data["price_4h"]
    p1d = data["price_1d"]
    dxy = data["dxy_1d"]

    trap = s2b_trap_analysis(trades["S2B-Hammer"], p30)
    checklists, tb_results = {}, {}
    for sid in SID_ORDER:
        print(f"    [{sid}] checklist + time_bleed...")
        checklists[sid] = entry_checklist(trades[sid], p30, p4h, p1d, dxy, sid)
        tb_results[sid]  = time_bleed_features(trades[sid], p30, p4h, p1d, dxy, sid)

    return {"trades": trades, "summs": summs, "trap": trap,
            "checklists": checklists, "tb_results": tb_results}


# ── DataFrame builders ────────────────────────────────────────────────────────

def _pct(v, dec=1):
    return f"{v*100:.{dec}f}%" if pd.notna(v) else "—"


def _build_overview(results: dict) -> pd.DataFrame:
    rows = []
    for sid in SID_ORDER:
        s  = results["summs"][sid]
        tb = results["tb_results"][sid]
        rows.append({
            "策略":            sid,
            "交易數":          s["total_trades"],
            "勝率":            _pct(s["win_rate"]),
            "獲利因子":        round(s["profit_factor"], 3),
            "淨盈虧 ($)":      round(s["net_pnl"], 0),
            "最大回撤 ($)":    round(s["max_drawdown"], 0),
            "time_bleed / 虧損": _pct(tb["time_bleed_pct_of_losses"]),
            "總虧損筆數":      tb["total_losses"],
            "time_bleed 筆數": tb["time_bleed_count"],
        })
    return pd.DataFrame(rows)


def _build_trap(trap: dict) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame | None]:
    """Returns (group_df, fail_type_df, fb_stats_df)."""
    hbs = trap["high_bb_summary"]
    lbs = trap["low_bb_summary"]
    grp = pd.DataFrame([
        {"分組": "BB%B >= 0.8（高位）", "交易數": hbs["total"],
         "勝率": _pct(hbs["win_rate"]), "敗筆數": hbs["losses"]},
        {"分組": "BB%B < 0.8（低位）",  "交易數": lbs["total"],
         "勝率": _pct(lbs["win_rate"]), "敗筆數": lbs["losses"]},
    ])

    fd = trap.get("fail_dist", {})
    fp = trap.get("fail_pct", {})
    fail_rows = [{"失敗類型": k, "筆數": v, "佔高位敗筆%": f"{fp.get(k, 0):.1f}%"}
                 for k, v in fd.items()]
    fail_df = pd.DataFrame(fail_rows) if fail_rows else pd.DataFrame()

    fb_df = None
    if trap.get("false_breakout_stats"):
        fb = trap["false_breakout_stats"]
        fb_df = pd.DataFrame([{
            "false_breakout 筆數":   fb["count"],
            "平均 MFE%":             _pct(fb["mean_mfe_pct"] / 100, 3),
            "中位 MFE%":             _pct(fb["median_mfe_pct"] / 100, 3),
            "P75 MFE%":              _pct(fb["p75_mfe_pct"] / 100, 3),
            "最大 MFE%":             _pct(fb["max_mfe_pct"] / 100, 3),
            "可存活（MFE<0.5%）":    fb["survivable_count"],
            "可存活%":               f"{fb['survivable_pct']:.1f}%",
        }])

    return grp, fail_df, fb_df


def _build_checklist(cl: dict) -> dict[str, pd.DataFrame]:
    out = {}
    if "by_4h_state" in cl and not cl["by_4h_state"].empty:
        out["by_4h_state"] = cl["by_4h_state"].reset_index()
    if "cross_4h_x_bb" in cl:
        wr = cl["cross_4h_x_bb"]["win_rate"].map(
            lambda v: _pct(v) if pd.notna(v) else "n<5")
        n  = cl["cross_4h_x_bb"]["counts"]
        out["4H_x_BB_勝率"] = wr.reset_index()
        out["4H_x_BB_筆數"] = n.reset_index()
    if "by_dxy" in cl and not cl["by_dxy"].empty:
        out["by_dxy"] = cl["by_dxy"].reset_index()
    return out


def _build_tb(tb: dict) -> dict[str, pd.DataFrame]:
    out = {}
    for key in ("by_4h_state", "by_session", "by_dxy"):
        if key in tb and not tb[key].empty:
            out[key] = tb[key].reset_index()
    return out


# ── Excel writer ──────────────────────────────────────────────────────────────

# Colour palette
_HDR_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
_SUB_FILL  = PatternFill("solid", fgColor="2E75B6")   # medium blue
_GREEN     = PatternFill("solid", fgColor="C6EFCE")
_YELLOW    = PatternFill("solid", fgColor="FFEB9C")
_RED       = PatternFill("solid", fgColor="FFC7CE")
_WHITE_FT  = Font(bold=True, color="FFFFFF")
_BOLD      = Font(bold=True)
_THIN      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _write_header(ws, row: int, col: int, text: str, span: int = 1, sub: bool = False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = _WHITE_FT
    cell.fill      = _SUB_FILL if sub else _HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)


def _write_section_title(ws, row: int, ncols: int, text: str):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = Font(bold=True, size=11)
    cell.fill      = PatternFill("solid", fgColor="D6DCE4")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=ncols)
    return row + 1


def _wr_fill(val_str: str):
    """Return fill based on win-rate string like '58.3%'."""
    if val_str == "—" or "n<5" in val_str:
        return None
    try:
        v = float(val_str.rstrip("%")) / 100
        if v >= 0.58: return _GREEN
        if v >= 0.45: return _YELLOW
        return _RED
    except Exception:
        return None


def _write_df(ws, df: pd.DataFrame, start_row: int, wr_cols: list[str] | None = None,
              title: str | None = None) -> int:
    """Write a DataFrame at start_row, return next available row."""
    if df is None or df.empty:
        return start_row

    r = start_row
    if title:
        r = _write_section_title(ws, r, len(df.columns), title)

    # Header row
    for ci, col in enumerate(df.columns, 1):
        _write_header(ws, r, ci, str(col), sub=True)
    r += 1

    # Data rows
    for _, row_data in df.iterrows():
        for ci, col in enumerate(df.columns, 1):
            val = row_data[col]
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border = _THIN
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if wr_cols and col in wr_cols:
                fill = _wr_fill(str(val))
                if fill:
                    cell.fill = fill
        r += 1

    # Auto-width
    for ci in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            len(str(df.columns[ci - 1])),
            *[len(str(df.iloc[ri, ci - 1])) for ri in range(len(df))],
        ) + 2
        ws.column_dimensions[col_letter].width = min(max_len, 25)

    return r + 1  # blank row after table


def _write_excel_file(results: dict, mode: str, out_path: Path):
    label = DATASET_MODES[mode]["label"]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        dummy = pd.DataFrame({"_": []})
        for sheet in ["策略概覽", "S2B垂頭陷阱", "S1進場清單", "S2A進場清單", "S2B進場清單", "TimeBleed"]:
            dummy.to_excel(writer, sheet_name=sheet, index=False)

    # Re-open for formatting
    wb = load_workbook(out_path)

    # ── 策略概覽 ──────────────────────────────────────────────────────
    ws = wb["策略概覽"]
    ws.delete_rows(1, ws.max_row)
    _write_section_title(ws, 1, 9, f"策略概覽 · {label}")
    df_ov = _build_overview(results)
    _write_df(ws, df_ov, 2, wr_cols=["勝率"])

    # ── S2B 垂頭陷阱 ─────────────────────────────────────────────────
    ws = wb["S2B垂頭陷阱"]
    ws.delete_rows(1, ws.max_row)
    r = _write_section_title(ws, 1, 4, f"S2B 垂頭陷阱（BB%B 分組）· {label}")
    grp_df, fail_df, fb_df = _build_trap(results["trap"])
    r = _write_df(ws, grp_df, r, wr_cols=["勝率"], title=None)
    r = _write_df(ws, fail_df, r, title="高位 BB 敗筆失敗類型分佈")
    if fb_df is not None:
        r = _write_df(ws, fb_df, r, title="False Breakout 統計（多單失敗曾反彈幅度）")

    # ── 進場清單 ─────────────────────────────────────────────────────
    for sid in SID_ORDER:
        short_name = sid.replace("-", "").replace("AweWithBB", "S1").replace("RSI", "S2A").replace("Hammer", "S2B")
        sheet = sid.split("-")[0] + sid.split("-")[1][:3] + "進場清單"
        # fix sheet name
        sheet_map = {"S1Awe進場清單": "S1進場清單", "S2ARSI進場清單": "S2A進場清單", "S2BHam進場清單": "S2B進場清單"}
        sheet = sheet_map.get(sheet, sheet)

        ws = wb[sheet]
        ws.delete_rows(1, ws.max_row)
        r = _write_section_title(ws, 1, 6, f"{sid} 進場清單 · {label}")

        cl_data = _build_checklist(results["checklists"][sid])
        if "by_4h_state" in cl_data:
            r = _write_df(ws, cl_data["by_4h_state"], r, wr_cols=["win_rate", "wins"], title="4H RSI 狀態 × 勝率")
        if "4H_x_BB_勝率" in cl_data:
            r = _write_df(ws, cl_data["4H_x_BB_勝率"], r, wr_cols=["BB<0.8", "BB>=0.8"], title="4H 狀態 × BB 位置 勝率矩陣")
        if "4H_x_BB_筆數" in cl_data:
            r = _write_df(ws, cl_data["4H_x_BB_筆數"], r, title="4H 狀態 × BB 位置 筆數")
        if "by_dxy" in cl_data:
            r = _write_df(ws, cl_data["by_dxy"], r, wr_cols=["win_rate"], title="DXY RSI Bucket × 勝率")

    # ── TimeBleed ─────────────────────────────────────────────────────
    ws = wb["TimeBleed"]
    ws.delete_rows(1, ws.max_row)
    r = _write_section_title(ws, 1, 8, f"Time-Bleed 特徵分析 · {label}")
    r += 1

    for sid in SID_ORDER:
        tb_data = _build_tb(results["tb_results"][sid])
        tb_sum = results["tb_results"][sid]
        r = _write_section_title(ws, r, 8,
            f"{sid}  |  time_bleed {tb_sum['time_bleed_count']}/{tb_sum['total_losses']} 虧損"
            f" = {_pct(tb_sum['time_bleed_pct_of_losses'])}")
        if "by_4h_state" in tb_data:
            r = _write_df(ws, tb_data["by_4h_state"], r, wr_cols=["win_rate"], title="by 4H 狀態")
        if "by_session" in tb_data:
            r = _write_df(ws, tb_data["by_session"], r, wr_cols=["win_rate"], title="by 交易時段")
        if "by_dxy" in tb_data:
            r = _write_df(ws, tb_data["by_dxy"], r, wr_cols=["win_rate"], title="by DXY Bucket")

    # Freeze top row on every sheet
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"  -> {out_path.name}")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    print("Loading price / DXY data (shared)...")
    data = _load_common_data()
    for k, v in data.items():
        if v is not None:
            print(f"  [{k}] {len(v)} bars")

    for mode in ["short", "long"]:
        label = DATASET_MODES[mode]["label"]
        print(f"\n[{mode.upper()}] {label}")
        results = _run_mode(mode, data)
        for sid in SID_ORDER:
            s = results["summs"][sid]
            print(f"  {sid}: {s['total_trades']} trades, WR={s['win_rate']:.1%}, PF={s['profit_factor']:.2f}")

        fname = f"deep_{mode}.xlsx"
        print(f"  Writing {fname}...")
        _write_excel_file(results, mode, OUT_DIR / fname)

    print(f"\nDone. Open both files side by side:")
    print(f"  {OUT_DIR / 'deep_short.xlsx'}")
    print(f"  {OUT_DIR / 'deep_long.xlsx'}")


if __name__ == "__main__":
    main()
